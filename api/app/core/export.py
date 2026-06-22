"""Utilitário genérico para geração de planilhas XLSX."""

from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

MAX_EXPORT_ROWS = 10_000


def generate_xlsx(
    *,
    title: str,
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Dados",
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            if isinstance(value, datetime):
                value = value.strftime("%d/%m/%Y %H:%M")
            elif isinstance(value, date):
                value = value.strftime("%d/%m/%Y")
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
