"""Testes dos endpoints de exportação Excel."""

import io

import pytest
from openpyxl import load_workbook

from tests.conftest import TENANT_A, auth_header, make_token

PREFIX = "/api/v1"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def auth_none() -> dict[str, str]:
    return {"Authorization": f"Bearer {make_token(TENANT_A, 1, [])}"}


def _load_xlsx(content: bytes):
    return load_workbook(io.BytesIO(content))


# ---------------------------------------------------------------------------
# Occurrences
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_occurrences_export_returns_xlsx(client):
    r = await client.get(
        f"{PREFIX}/occurrences/export", headers=auth_header(TENANT_A)
    )
    assert r.status_code == 200
    assert XLSX_MIME in r.headers["content-type"]
    wb = _load_xlsx(r.content)
    ws = wb.active
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 2).value == "Título"


@pytest.mark.asyncio
async def test_occurrences_export_forbidden(client):
    r = await client.get(
        f"{PREFIX}/occurrences/export", headers=auth_none()
    )
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_occurrences_export_with_data(client):
    await client.post(
        f"{PREFIX}/occurrences",
        headers=auth_header(TENANT_A),
        json={"title": "Export test", "category": "general"},
    )
    r = await client.get(
        f"{PREFIX}/occurrences/export", headers=auth_header(TENANT_A)
    )
    assert r.status_code == 200
    wb = _load_xlsx(r.content)
    ws = wb.active
    assert ws.max_row >= 2


# ---------------------------------------------------------------------------
# Maintenance
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_maintenance_export_returns_xlsx(client):
    r = await client.get(
        f"{PREFIX}/maintenance/export", headers=auth_header(TENANT_A)
    )
    assert r.status_code == 200
    assert XLSX_MIME in r.headers["content-type"]
    wb = _load_xlsx(r.content)
    ws = wb.active
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 5).value == "Status"


@pytest.mark.asyncio
async def test_maintenance_export_forbidden(client):
    r = await client.get(
        f"{PREFIX}/maintenance/export", headers=auth_none()
    )
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# Checklists executions
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_checklists_export_returns_xlsx(client):
    r = await client.get(
        f"{PREFIX}/checklists/executions/export",
        headers=auth_header(TENANT_A),
    )
    assert r.status_code == 200
    assert XLSX_MIME in r.headers["content-type"]
    wb = _load_xlsx(r.content)
    ws = wb.active
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 2).value == "Template"


@pytest.mark.asyncio
async def test_checklists_export_forbidden(client):
    r = await client.get(
        f"{PREFIX}/checklists/executions/export",
        headers=auth_none(),
    )
    assert r.status_code == 403


# ---------------------------------------------------------------------------
# Registries
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_registries_export_returns_xlsx(client):
    r = await client.get(
        f"{PREFIX}/registries/export", headers=auth_header(TENANT_A)
    )
    assert r.status_code == 200
    assert XLSX_MIME in r.headers["content-type"]
    wb = _load_xlsx(r.content)
    ws = wb.active
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 3).value == "Categoria"


@pytest.mark.asyncio
async def test_registries_export_forbidden(client):
    r = await client.get(
        f"{PREFIX}/registries/export", headers=auth_none()
    )
    assert r.status_code == 403
